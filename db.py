# -*- coding: utf-8 -*-
"""
db.py
Capa de acceso a la base de datos simulada (Excel). Se separa de bot.py
para mantener la lógica de persistencia desacoplada de la lógica de
conversación, y para poder manejar en un solo lugar los errores típicos
de trabajar con un archivo Excel (bloqueado por estar abierto, corrupto,
inexistente, etc.).
"""
import time
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from config import DB_PATH

logger = logging.getLogger(__name__)

REINTENTOS = 3
ESPERA_ENTRE_REINTENTOS = 0.4  # segundos


class DBError(Exception):
    """Error de acceso a la base de datos simulada."""


def _con_reintentos(func, *args, **kwargs):
    """Reintenta una operación de archivo ante errores de bloqueo (PermissionError),
    típico cuando el Excel está abierto en otro programa al mismo tiempo."""
    ultimo_error = None
    for intento in range(1, REINTENTOS + 1):
        try:
            return func(*args, **kwargs)
        except PermissionError as e:
            ultimo_error = e
            logger.warning("Archivo de base de datos bloqueado (intento %s/%s)",
                            intento, REINTENTOS)
            time.sleep(ESPERA_ENTRE_REINTENTOS)
        except FileNotFoundError as e:
            raise DBError(
                f"No se encontró '{DB_PATH}'. Corré crear_db.py primero."
            ) from e
    raise DBError(
        f"No se pudo acceder a '{DB_PATH}' tras {REINTENTOS} intentos. "
        f"¿Está abierto en Excel? Cerralo e intentá de nuevo."
    ) from ultimo_error


def crear_base():
    """Inicializa el archivo de base de datos con sus dos hojas."""
    wb = openpyxl.Workbook()

    ws_emp = wb.active
    ws_emp.title = "empleados"
    headers = ["legajo", "nombre", "saldo_dias", "jefe_directo"]
    ws_emp.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws_emp.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    empleados = [
        (1001, "Guadalupe Soraire", 14, "Mario López"),
        (1002, "Luca Vaccaro", 8, "Mario López"),
        (1003, "Rosario Mallón", 20, "Andrea Ramos"),
        (1004, "Mauro Rodríguez", 3, "Andrea Ramos"),
    ]
    for fila in empleados:
        ws_emp.append(fila)

    ws_sol = wb.create_sheet("solicitudes")
    headers_sol = ["id_solicitud", "legajo", "fecha_inicio", "fecha_fin",
                   "dias_solicitados", "estado", "fecha_registro", "resuelto_por"]
    ws_sol.append(headers_sol)
    for col in range(1, len(headers_sol) + 1):
        c = ws_sol.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    wb.save(DB_PATH)
    logger.info("Base de datos '%s' creada.", DB_PATH)


def buscar_empleado(legajo: int):
    """Devuelve dict con datos del empleado, o None si no existe."""
    def _op():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb["empleados"]
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=1).value == legajo:
                return {
                    "fila": fila,
                    "legajo": legajo,
                    "nombre": ws.cell(row=fila, column=2).value,
                    "saldo": ws.cell(row=fila, column=3).value,
                    "jefe": ws.cell(row=fila, column=4).value,
                }
        return None
    return _con_reintentos(_op)


def actualizar_saldo(legajo: int, nuevo_saldo: int):
    def _op():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb["empleados"]
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=1).value == legajo:
                ws.cell(row=fila, column=3).value = nuevo_saldo
                wb.save(DB_PATH)
                return True
        return False
    return _con_reintentos(_op)


def registrar_solicitud(legajo, inicio, fin, dias, estado):
    """Inserta una nueva solicitud y devuelve el id_solicitud generado."""
    def _op():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb["solicitudes"]
        nuevo_id = ws.max_row  # max_row=1 (solo header) -> primer id = 1
        ws.append([nuevo_id, legajo,
                   inicio.strftime("%d/%m/%Y"), fin.strftime("%d/%m/%Y"),
                   dias, estado, datetime.now().strftime("%Y-%m-%d %H:%M"), ""])
        wb.save(DB_PATH)
        return nuevo_id
    return _con_reintentos(_op)


def actualizar_estado_solicitud(id_solicitud: int, nuevo_estado: str, resuelto_por: str = ""):
    def _op():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb["solicitudes"]
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=1).value == id_solicitud:
                ws.cell(row=fila, column=6).value = nuevo_estado
                ws.cell(row=fila, column=8).value = resuelto_por
                legajo = ws.cell(row=fila, column=2).value
                dias = ws.cell(row=fila, column=5).value
                wb.save(DB_PATH)
                return legajo, dias
        return None
    return _con_reintentos(_op)


def obtener_solicitud(id_solicitud: int):
    def _op():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb["solicitudes"]
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=1).value == id_solicitud:
                return {
                    "id": id_solicitud,
                    "legajo": ws.cell(row=fila, column=2).value,
                    "fecha_inicio": ws.cell(row=fila, column=3).value,
                    "fecha_fin": ws.cell(row=fila, column=4).value,
                    "dias": ws.cell(row=fila, column=5).value,
                    "estado": ws.cell(row=fila, column=6).value,
                }
        return None
    return _con_reintentos(_op)


def historial_por_legajo(legajo: int):
    def _op():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb["solicitudes"]
        resultado = []
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=2).value == legajo:
                resultado.append({
                    "id": ws.cell(row=fila, column=1).value,
                    "fecha_inicio": ws.cell(row=fila, column=3).value,
                    "fecha_fin": ws.cell(row=fila, column=4).value,
                    "dias": ws.cell(row=fila, column=5).value,
                    "estado": ws.cell(row=fila, column=6).value,
                })
        return resultado
    return _con_reintentos(_op)


def solicitudes_pendientes():
    def _op():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb["solicitudes"]
        resultado = []
        for fila in range(2, ws.max_row + 1):
            if ws.cell(row=fila, column=6).value == "PENDIENTE_APROBACION":
                resultado.append({
                    "id": ws.cell(row=fila, column=1).value,
                    "legajo": ws.cell(row=fila, column=2).value,
                    "fecha_inicio": ws.cell(row=fila, column=3).value,
                    "fecha_fin": ws.cell(row=fila, column=4).value,
                    "dias": ws.cell(row=fila, column=5).value,
                })
        return resultado
    return _con_reintentos(_op)
